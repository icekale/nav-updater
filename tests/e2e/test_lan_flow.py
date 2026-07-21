import re
from datetime import date
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.catalog import private_product_code
from app.config import Settings
from app.db import Base
from app.jobs.service import batch_manage_runs
from app.main import create_app
from app.models import (
    AuditLog,
    Meeting,
    OcrRegressionSample,
    OcrReviewSample,
    Product,
    RunFile,
    RunItem,
    UpdateRun,
    User,
)


def meeting_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "近期会议更新"
    sheet.append(["近期资本市场相关会议更新"])
    sheet.append(
        [
            "会议/事件",
            "日期",
            "性质/层级",
            "核心表述",
            "资本市场影响",
            "投研映射",
            "后续跟踪",
            "来源链接",
            "更新时间",
        ]
    )
    sheet.append(
        [
            "2026陆家嘴论坛",
            "2026-06-17至2026-06-18",
            "金融高层论坛",
            "服务高质量发展",
            "投融资综合改革",
            "科技成长",
            "跟踪改革细则",
            "https://example.test/source",
            "2026-07-18",
        ]
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_login_page_is_available() -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=Path("/tmp/nav-updater-test"),
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )
    client = TestClient(create_app(settings=settings, session_factory=factory))
    response = client.get("/login")
    assert response.status_code == 200
    assert "登录" in response.text


def test_user_history_foreign_keys_use_set_null() -> None:
    assert next(
        foreign_key.ondelete
        for foreign_key in UpdateRun.__table__.foreign_keys
        if foreign_key.parent.name == "operator_id"
    ) == "SET NULL"
    assert next(
        foreign_key.ondelete
        for foreign_key in AuditLog.__table__.foreign_keys
        if foreign_key.parent.name == "actor_id"
    ) == "SET NULL"
    assert UpdateRun.__table__.c.operator_id.nullable is True


def _test_app(tmp_path: Path):
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )
    return create_app(settings=settings, session_factory=factory), factory


def _token(client: TestClient, path: str) -> str:
    response = client.get(path)
    match = re.search(r'name="token" value="([^"]+)"', response.text)
    assert match is not None
    return match.group(1)


def _login_as_admin(client: TestClient) -> None:
    response = client.post(
        "/login",
        data={"username": "admin", "password": "change-me", "token": _token(client, "/login")},
        follow_redirects=False,
    )
    assert response.status_code == 303


def _create_run_with_artifacts(
    factory: sessionmaker,
    data_dir: Path,
    *,
    status: str = "completed",
    directory: str = "run-artifacts",
) -> tuple[int, int, Path, Path]:
    session = factory()
    try:
        admin = session.query(User).filter_by(username="admin").one()
        run_dir = data_dir / "runs" / directory
        run_dir.mkdir(parents=True)
        workbook = run_dir / "input.xlsx"
        image = run_dir / "source.png"
        result = run_dir / "result.xlsx"
        workbook.write_bytes(b"input")
        image.write_bytes(b"image")
        result.write_bytes(b"result")
        run = UpdateRun(
            operator_id=admin.id,
            cutoff_date=date(2026, 7, 17),
            status=status,
            output_path=str(result),
        )
        session.add(run)
        session.flush()
        item = RunItem(
            run_id=run.id,
            excel_row=2,
            original_values={"product_name": "产品A"},
        )
        session.add_all(
            [
                RunFile(
                    run_id=run.id,
                    file_type="workbook",
                    original_name="input.xlsx",
                    storage_path=str(workbook),
                    sha256="0" * 64,
                ),
                RunFile(
                    run_id=run.id,
                    file_type="image",
                    original_name="source.png",
                    storage_path=str(image),
                    sha256="1" * 64,
                ),
                item,
            ]
        )
        session.flush()
        session.add_all(
            [
                AuditLog(
                    actor_id=admin.id,
                    action="create",
                    object_type="update_run",
                    object_id=str(run.id),
                ),
                AuditLog(
                    actor_id=admin.id,
                    action="manual_review",
                    object_type="run_item",
                    object_id=str(item.id),
                ),
            ]
        )
        session.commit()
        return run.id, item.id, run_dir, result
    finally:
        session.close()


def test_run_deletion_cascades_ocr_review_samples(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app):
        run_id, item_id, _, _ = _create_run_with_artifacts(factory, tmp_path)
        session = factory()
        try:
            sample = OcrReviewSample(
                run_id=run_id,
                run_item_id=item_id,
                excel_product_name="产品A",
                review_version=1,
                ocr_match_source="image",
                ocr_metric_values={"weekly": "0.01"},
                ocr_metric_status={"weekly": "extracted"},
                confirmed_metric_values={"weekly": "0.01"},
                confirmed_metric_status={"weekly": "manual"},
                review_note="人工确认",
            )
            session.add(sample)
            session.commit()
            run = session.get(UpdateRun, run_id)
            assert run is not None
            session.delete(run)
            session.commit()
            assert session.query(OcrReviewSample).count() == 0
        finally:
            session.close()


def test_regression_sample_survives_run_delete(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        run_id, item_id, _, _ = _create_run_with_artifacts(factory, tmp_path)
        sample_path = tmp_path / "ocr-quality" / "samples" / "sample.png"
        sample_path.parent.mkdir(parents=True)
        sample_path.write_bytes(b"sample")
        session = factory()
        try:
            session.add(
                OcrRegressionSample(
                    image_path=str(sample_path),
                    image_sha256="a" * 64,
                    source_run_id=run_id,
                    source_item_id=item_id,
                    source_label="管理员复核案例",
                    excel_product_name="产品A",
                    candidate_names=["产品A"],
                    expected_metric_values={"mtd": "-0.0633"},
                    expected_metric_status={"mtd": "extracted"},
                    note="确认",
                    is_active=True,
                )
            )
            session.commit()
        finally:
            session.close()
        deleted = client.post(
            f"/updates/{run_id}/delete",
            data={"token": _token(client, "/updates")},
            follow_redirects=False,
        )

    assert deleted.status_code == 303
    session = factory()
    try:
        sample = session.query(OcrRegressionSample).one()
        assert sample.source_run_id is None
        assert sample.source_item_id is None
        assert Path(sample.image_path).exists()
    finally:
        session.close()


def test_batch_requeue_moves_each_completed_run_back_to_the_queue(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        first_id, _, _, _ = _create_run_with_artifacts(factory, tmp_path, directory="batch-a")
        second_id, _, _, _ = _create_run_with_artifacts(factory, tmp_path, directory="batch-b")

        response = client.post(
            "/updates/batch",
            data={
                "token": _token(client, "/updates"),
                "action": "requeue",
                "run_ids": [str(first_id), str(second_id)],
            },
            follow_redirects=False,
        )

    assert response.status_code == 303
    session = factory()
    try:
        assert session.get(UpdateRun, first_id).status == "uploaded"
        assert session.get(UpdateRun, second_id).status == "uploaded"
        assert (
            session.query(AuditLog).filter_by(action="queue", object_type="update_run").count() == 2
        )
    finally:
        session.close()


def test_batch_delete_preserves_processing_run_and_deletes_completed_runs(
    tmp_path: Path,
) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        completed_id, _, completed_dir, _ = _create_run_with_artifacts(
            factory, tmp_path, directory="batch-delete-a"
        )
        second_completed_id, _, second_completed_dir, _ = _create_run_with_artifacts(
            factory, tmp_path, directory="batch-delete-b"
        )
        processing_id, _, processing_dir, _ = _create_run_with_artifacts(
            factory, tmp_path, status="processing", directory="batch-processing"
        )

        response = client.post(
            "/updates/batch",
            data={
                "token": _token(client, "/updates"),
                "action": "delete",
                "run_ids": [str(completed_id), str(second_completed_id), str(processing_id)],
            },
            follow_redirects=False,
        )
        history = client.get(response.headers["location"])

    assert response.status_code == 303
    assert "已删除 2 个批次，跳过处理中 1 个" in history.text
    session = factory()
    try:
        assert session.get(UpdateRun, completed_id) is None
        assert session.get(UpdateRun, second_completed_id) is None
        assert session.get(UpdateRun, processing_id) is not None
        assert session.query(AuditLog).filter_by(
            action="delete", object_type="update_run", object_id=str(completed_id)
        ).count() == 1
        assert session.query(AuditLog).filter_by(
            action="delete", object_type="update_run", object_id=str(second_completed_id)
        ).count() == 1
    finally:
        session.close()
    assert not completed_dir.exists()
    assert not second_completed_dir.exists()
    assert processing_dir.exists()


def test_batch_rejects_empty_selection_and_unknown_action(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        run_id, _, _, _ = _create_run_with_artifacts(factory, tmp_path, directory="batch-input")

        empty = client.post(
            "/updates/batch",
            data={"token": _token(client, "/updates"), "action": "requeue"},
            follow_redirects=False,
        )
        empty_history = client.get(empty.headers["location"])
        unknown = client.post(
            "/updates/batch",
            data={
                "token": _token(client, "/updates"),
                "action": "unknown",
                "run_ids": str(run_id),
            },
            follow_redirects=False,
        )
        unknown_history = client.get(unknown.headers["location"])

    assert empty.status_code == 303
    assert "请选择至少一个批次" in empty_history.text
    assert unknown.status_code == 303
    assert "批量操作无效" in unknown_history.text
    session = factory()
    try:
        assert session.get(UpdateRun, run_id).status == "completed"
    finally:
        session.close()


def test_batch_rejects_runs_outside_the_current_history_page(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        run_ids = [
            _create_run_with_artifacts(factory, tmp_path, directory=f"batch-limit-{index}")[0]
            for index in range(51)
        ]

        response = client.post(
            "/updates/batch",
            data={
                "token": _token(client, "/updates"),
                "action": "delete",
                "run_ids": [str(run_id) for run_id in run_ids],
            },
            follow_redirects=False,
        )
        history = client.get(response.headers["location"])

    assert response.status_code == 303
    assert "只能操作当前页显示的批次" in history.text
    session = factory()
    try:
        assert session.get(UpdateRun, run_ids[0]).status == "completed"
        assert session.get(UpdateRun, run_ids[-1]).status == "completed"
    finally:
        session.close()


def test_batch_delete_rechecks_a_stale_run_status_under_lock(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app):
        run_id, _, run_dir, _ = _create_run_with_artifacts(
            factory, tmp_path, directory="batch-stale-processing"
        )
        session = factory()
        concurrent_session = factory()
        try:
            assert session.get(UpdateRun, run_id).status == "completed"
            concurrent_session.get(UpdateRun, run_id).status = "processing"
            concurrent_session.commit()

            result = batch_manage_runs(
                session,
                [run_id],
                action="delete",
                data_dir=tmp_path,
                actor_id=1,
            )

            assert result.skipped_processing == 1
            assert result.deleted == 0
        finally:
            session.close()
            concurrent_session.close()

    session = factory()
    try:
        assert session.get(UpdateRun, run_id) is not None
    finally:
        session.close()
    assert run_dir.exists()


def test_batch_requeue_does_not_commit_without_its_audit_log(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app):
        run_id, _, _, _ = _create_run_with_artifacts(
            factory, tmp_path, directory="batch-atomic-requeue"
        )
        session = factory()
        try:
            @event.listens_for(session, "before_commit")
            def reject_queue_audit(active_session) -> None:
                if any(
                    isinstance(instance, AuditLog) and instance.action == "queue"
                    for instance in active_session.new
                ):
                    raise RuntimeError("audit write failed")

            with pytest.raises(RuntimeError, match="audit write failed"):
                batch_manage_runs(
                    session,
                    [run_id],
                    action="requeue",
                    data_dir=tmp_path,
                    actor_id=1,
                )
            session.rollback()
        finally:
            session.close()

    session = factory()
    try:
        assert session.get(UpdateRun, run_id).status == "completed"
        assert (
            session.query(AuditLog)
            .filter_by(action="queue", object_type="update_run", object_id=str(run_id))
            .count()
            == 0
        )
    finally:
        session.close()


def test_review_sample_is_atomic_when_sample_write_fails(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app, raise_server_exceptions=False) as client:
        _login_as_admin(client)
        run_id, item_id, _, _ = _create_run_with_artifacts(factory, tmp_path)
        token = _token(client, f"/updates/{run_id}/review")

        def reject_sample(active_session) -> None:
            if active_session.query(OcrReviewSample).count():
                raise SQLAlchemyError("sample write failed")

        event.listen(factory.class_, "before_commit", reject_sample)
        try:
            response = client.post(
                f"/updates/{run_id}/items/{item_id}/review",
                data={
                    "token": token,
                    "product_choice": "create_private",
                    "review_note": "以管理人净值表为准",
                    "weekly": "1.23",
                },
            )
        finally:
            event.remove(factory.class_, "before_commit", reject_sample)

    assert response.status_code == 500
    assert "保存审核失败，请重试" in response.text
    session = factory()
    try:
        item = session.get(RunItem, item_id)
        assert item is not None
        assert item.match_source != "manual"
        assert session.query(OcrReviewSample).filter_by(run_item_id=item_id).count() == 0
    finally:
        session.close()


def test_quality_center_renders_metrics_and_review_links(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        run_id, item_id, _, _ = _create_run_with_artifacts(factory, tmp_path)
        session = factory()
        try:
            product = Product(
                product_name="质检产品", product_code="QUALITY-001", product_type="private"
            )
            session.add(product)
            session.flush()
            item = session.get(RunItem, item_id)
            assert item is not None
            item.row_status = "ready"
            item.metric_status = {"mtd": "source_blank"}
            session.add(
                OcrReviewSample(
                    run_id=run_id,
                    run_item_id=item_id,
                    product_id=product.id,
                    excel_product_name="质检产品",
                    review_version=1,
                    ocr_match_source="image",
                    ocr_product_id=product.id,
                    ocr_metric_values={},
                    ocr_metric_status={"weekly": "stale"},
                    confirmed_metric_values={"weekly": "0.01"},
                    confirmed_metric_status={"weekly": "manual"},
                    review_note="人工确认",
                )
            )
            session.commit()
        finally:
            session.close()

        response = client.get("/quality")

    assert response.status_code == 200
    assert "质检中心" in response.text
    assert "字段一致率" in response.text
    assert "漏识别" in response.text
    assert "source_blank" not in response.text
    assert f'/updates/{run_id}/review?show_all=1#review-item-{item_id}' in response.text


def test_private_product_monitoring_requires_login_renders_and_exports(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        page_without_login = client.get("/monitoring", follow_redirects=False)
        export_without_login = client.get("/monitoring/export.xlsx", follow_redirects=False)
        assert page_without_login.status_code == 303
        assert page_without_login.headers["location"] == "/login"
        assert export_without_login.status_code == 303
        assert export_without_login.headers["location"] == "/login"

        _login_as_admin(client)
        session = factory()
        try:
            admin = session.query(User).filter_by(username="admin").one()
            product = Product(
                product_name="监控产品", product_code="MONITOR-001", product_type="private"
            )
            session.add(product)
            session.flush()
            run = UpdateRun(
                operator_id=admin.id,
                cutoff_date=date(2026, 7, 20),
                status="completed_with_warnings",
            )
            session.add(run)
            session.flush()
            item = RunItem(
                run_id=run.id,
                product_id=product.id,
                excel_row=2,
                row_status="stale",
                metric_status={"weekly": "stale"},
            )
            session.add(item)
            session.flush()
            run_id = run.id
            item_id = item.id
            session.commit()
        finally:
            session.close()

        response = client.get("/monitoring?status=missing_data")
        exported = client.get("/monitoring/export.xlsx?status=missing_data")

    assert response.status_code == 200
    assert "产品监控" in response.text
    assert "数据缺失" in response.text
    assert "导出当前清单" in response.text
    assert f'/updates/{run_id}/review?show_all=1#review-item-{item_id}' in response.text
    assert exported.status_code == 200
    assert (
        exported.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=" in exported.headers["content-disposition"]
    assert "私募产品监控" in unquote(exported.headers["content-disposition"])
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    assert workbook.active.max_row == 2


def test_history_page_renders_batch_controls_for_visible_runs(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        first_id, _, _, _ = _create_run_with_artifacts(factory, tmp_path, directory="batch-ui-a")
        second_id, _, _, _ = _create_run_with_artifacts(factory, tmp_path, directory="batch-ui-b")
        response = client.get("/updates")
        stylesheet = client.get("/static/app.css")

    assert 'id="batch-run-form"' in response.text
    assert 'id="select-all-runs"' in response.text
    assert f'name="run_ids" value="{first_id}"' in response.text
    assert f'name="run_ids" value="{second_id}"' in response.text
    assert 'name="action" value="requeue"' in response.text
    assert 'name="action" value="delete"' in response.text
    assert "确定永久删除所选批次及其上传文件和结果文件吗？" in response.text
    assert ".batch-toolbar" in stylesheet.text
    assert ".batch-toolbar[hidden] { display: none; }" in stylesheet.text


def test_deleting_completed_run_removes_artifacts_and_old_audit_logs(
    tmp_path: Path,
) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        run_id, item_id, run_dir, result = _create_run_with_artifacts(factory, tmp_path)

        deleted = client.post(
            f"/updates/{run_id}/delete",
            data={"token": _token(client, "/updates")},
            follow_redirects=False,
        )

    assert deleted.status_code == 303
    assert deleted.headers["location"].startswith("/updates?notice=")
    session = factory()
    try:
        assert session.get(UpdateRun, run_id) is None
        assert session.query(RunFile).filter_by(run_id=run_id).count() == 0
        assert session.query(RunItem).filter_by(run_id=run_id).count() == 0
        deleted_log = session.query(AuditLog).filter_by(
            action="delete", object_type="update_run", object_id=str(run_id)
        ).one()
        assert deleted_log.context == {"deleted_item_count": 1, "deleted_file_count": 3}
        assert session.query(AuditLog).filter_by(
            object_type="run_item", object_id=str(item_id)
        ).count() == 0
    finally:
        session.close()
    assert not run_dir.exists()
    assert not result.exists()


def test_deleting_processing_run_is_rejected_without_removing_files(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        run_id, _, run_dir, result = _create_run_with_artifacts(
            factory, tmp_path, status="processing"
        )

        rejected = client.post(
            f"/updates/{run_id}/delete", data={"token": _token(client, "/updates")}
        )

    assert rejected.status_code == 409
    session = factory()
    try:
        assert session.get(UpdateRun, run_id) is not None
    finally:
        session.close()
    assert run_dir.exists()
    assert result.exists()


def test_deleting_run_keeps_a_file_outside_the_data_directory(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    outside_path = tmp_path.parent / "keep-me.xlsx"
    outside_path.write_bytes(b"keep")
    with TestClient(app) as client:
        _login_as_admin(client)
        run_id, _, _, _ = _create_run_with_artifacts(factory, tmp_path)
        session = factory()
        try:
            session.add(
                RunFile(
                    run_id=run_id,
                    file_type="workbook",
                    original_name="keep-me.xlsx",
                    storage_path=str(outside_path),
                    sha256="2" * 64,
                )
            )
            session.commit()
        finally:
            session.close()

        deleted = client.post(
            f"/updates/{run_id}/delete",
            data={"token": _token(client, "/updates")},
            follow_redirects=False,
        )

    assert deleted.status_code == 303
    assert outside_path.read_bytes() == b"keep"


def test_admin_deletes_other_user_and_retains_run_history(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        session = factory()
        try:
            operator = User(username="operator", password_hash="hash", role="user")
            session.add(operator)
            session.flush()
            run = UpdateRun(operator_id=operator.id, cutoff_date=date(2026, 7, 17))
            session.add(run)
            session.flush()
            session.add(
                AuditLog(
                    actor_id=operator.id,
                    action="create",
                    object_type="update_run",
                    object_id=str(run.id),
                )
            )
            session.commit()
            operator_id, run_id = operator.id, run.id
        finally:
            session.close()

        deleted = client.post(
            f"/admin/users/{operator_id}/delete",
            data={"token": _token(client, "/admin/users")},
            follow_redirects=False,
        )

    assert deleted.status_code == 303
    session = factory()
    try:
        assert session.get(User, operator_id) is None
        assert session.get(UpdateRun, run_id).operator_id is None
        assert session.query(AuditLog).filter_by(
            action="create", object_type="update_run", object_id=str(run_id)
        ).one().actor_id is None
        assert session.query(AuditLog).filter_by(
            action="delete", object_type="user", object_id=str(operator_id)
        ).count() == 1
    finally:
        session.close()


def test_admin_cannot_delete_self_or_the_last_admin(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        session = factory()
        try:
            admin_id = session.query(User).filter_by(username="admin").one().id
        finally:
            session.close()

        rejected = client.post(
            f"/admin/users/{admin_id}/delete", data={"token": _token(client, "/admin/users")}
        )

    assert rejected.status_code == 409
    session = factory()
    try:
        assert session.get(User, admin_id) is not None
    finally:
        session.close()


def test_preview_counts_values_and_confirmed_source_blanks(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        session = factory()
        try:
            admin = session.query(User).filter_by(username="admin").one()
            run = UpdateRun(
                operator_id=admin.id,
                cutoff_date=date(2026, 7, 17),
                status="completed",
            )
            session.add(run)
            session.flush()
            extracted_metrics = {
                "weekly",
                "mtd",
                "ytd",
                "annual_2019",
                "annual_2020",
                "annual_2021",
                "annual_2022",
                "annual_2024",
            }
            statuses = {metric: "extracted" for metric in extracted_metrics}
            statuses.update(
                {
                    "annual_2023": "stale",
                    "annual_2025": "source_blank",
                    "sharpe": "source_blank",
                    "max_drawdown": "source_blank",
                }
            )
            session.add(
                RunItem(
                    run_id=run.id,
                    excel_row=2,
                    row_status="partial",
                    metric_values={metric: "0.01" for metric in extracted_metrics},
                    metric_status=statuses,
                    error_reason="本次未识别：annual_2023；OCR 置信度较低",
                    original_values={"product_name": "浑瑾产品"},
                )
            )
            session.commit()
            run_id = run.id
        finally:
            session.close()

        preview = client.get(f"/updates/{run_id}/preview")

    assert "已确认 11 / 12 项（8 数值＋3 空值）" in preview.text
    assert "2023（%）" in preview.text
    assert "annual_2023" not in preview.text
    assert "待人工审核 0 条" in preview.text
    assert "可直接生成 1 条" in preview.text


def test_workspace_pages_render_sidebar_and_safe_delete_controls(tmp_path: Path) -> None:
    app, factory = _test_app(tmp_path)
    with TestClient(app) as client:
        _login_as_admin(client)
        run_id, _, _, _ = _create_run_with_artifacts(factory, tmp_path)
        session = factory()
        try:
            operator = User(username="operator", password_hash="hash", role="user")
            session.add(operator)
            session.commit()
            operator_id = operator.id
        finally:
            session.close()

        updates = client.get("/updates")
        users = client.get("/admin/users")
        stylesheet = client.get("/static/app.css")

    assert 'class="app-shell"' in updates.text
    assert 'class="app-sidebar"' in updates.text
    assert 'class="app-main"' in updates.text
    assert 'class="nav-link is-active" href="/updates" aria-current="page"' in updates.text
    assert f'action="/updates/{run_id}/delete"' in updates.text
    assert "确定永久删除此批次及其上传文件和结果文件吗？" in updates.text
    assert "创建账号" in updates.text
    assert f'action="/admin/users/{operator_id}/delete"' in users.text
    assert "当前登录账号，不能删除" in users.text
    assert ".app-sidebar" in stylesheet.text
    assert "width: 232px" in stylesheet.text
    assert "@media (max-width: 799px)" in stylesheet.text


def test_login_catalog_upload_and_queue_run(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )
    app = create_app(settings=settings, session_factory=factory)
    with TestClient(app) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        logged_in = client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        assert logged_in.status_code == 303

        catalog_page = client.get("/catalog")
        token = re.search(r'name="token" value="([^"]+)"', catalog_page.text).group(1)
        catalog_response = client.post(
            "/catalog/import",
            data={"token": token},
            files={
                "catalog_file": (
                    "catalog.csv",
                    "product_name,product_code,product_type\n"
                    "仁桥金选泽源5B,P001,private\n"
                    "浑瑾岳桐金选1号B,P002,private\n"
                    "开思金选港股通1号B,P003,private\n"
                    "静瑞金选价值灵动1号B,P004,private\n"
                    "勤辰金选创赢成长1号B,P005,private\n"
                    "易方达环保主题灵活配置混合A,P006,private\n",
                    "text/csv",
                )
            },
            follow_redirects=False,
        )
        assert catalog_response.status_code == 303
        assert "仁桥金选泽源5B" in client.get("/catalog").text

        new_page = client.get("/updates/new")
        assert "同一报告日期" in new_page.text
        assert "不同周度请分别新建批次" in new_page.text
        token = re.search(r'name="token" value="([^"]+)"', new_page.text).group(1)
        workbook = Path("tests/fixtures/net_value_template.xlsx").read_bytes()
        created = client.post(
            "/updates/new",
            data={"token": token, "cutoff_date": "2026-07-17"},
            files={
                "workbook": (
                    "template.xlsx",
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=False,
        )
        assert created.status_code == 303
        preview = client.get(created.headers["location"])
        assert preview.status_code == 200
        token = re.search(r'name="token" value="([^"]+)"', preview.text).group(1)
        processed = client.post(
            created.headers["location"].replace("/preview", "/process"),
            data={"token": token},
            follow_redirects=False,
        )
        assert processed.status_code == 303
        final_page = client.get(created.headers["location"])
        assert "已进入后台处理队列" in final_page.text
        downloaded = client.get(created.headers["location"].replace("/preview", "/download"))
        assert downloaded.status_code == 404


def test_upload_keeps_same_named_images_separate(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )

    with TestClient(create_app(settings=settings, session_factory=factory)) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        new_page = client.get("/updates/new")
        token = re.search(r'name="token" value="([^"]+)"', new_page.text).group(1)
        created = client.post(
            "/updates/new",
            data={"token": token, "cutoff_date": "2026-07-17"},
            files=[
                (
                    "workbook",
                    (
                        "template.xlsx",
                        Path("tests/fixtures/net_value_template.xlsx").read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                ),
                ("images", ("snapshot.png", b"first image", "image/png")),
                ("images", ("snapshot.png", b"second image", "image/png")),
            ],
            follow_redirects=False,
        )
        run_id = int(re.search(r"/updates/(\d+)/preview", created.headers["location"]).group(1))

    session = factory()
    try:
        images = session.query(RunFile).filter_by(run_id=run_id, file_type="image").all()
        assert len(images) == 2
        assert len({image.storage_path for image in images}) == 2
        assert {image.original_name for image in images} == {"snapshot.png"}
        assert {Path(image.storage_path).read_bytes() for image in images} == {
            b"first image",
            b"second image",
        }
    finally:
        session.close()


def test_regenerate_requeues_completed_run_for_worker(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )

    with TestClient(create_app(settings=settings, session_factory=factory)) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        session = factory()
        try:
            admin = session.query(User).filter_by(username="admin").one()
            run = UpdateRun(
                operator_id=admin.id,
                cutoff_date=date(2026, 7, 17),
                status="completed",
                output_path=str(tmp_path / "old-result.xlsx"),
            )
            session.add(run)
            session.commit()
            run_id = run.id
        finally:
            session.close()

        preview = client.get(f"/updates/{run_id}/preview")
        token = re.search(r'name="token" value="([^"]+)"', preview.text).group(1)
        requested = client.post(
            f"/updates/{run_id}/process",
            data={"token": token},
            follow_redirects=False,
        )
        assert requested.status_code == 303

    session = factory()
    try:
        run = session.get(UpdateRun, run_id)
        assert run.status == "uploaded"
        assert run.output_path is None
    finally:
        session.close()


def test_manual_review_is_rejected_while_run_is_processing(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )

    with TestClient(create_app(settings=settings, session_factory=factory)) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        session = factory()
        try:
            admin = session.query(User).filter_by(username="admin").one()
            product = Product(product_name="产品A", product_code="P001", product_type="private")
            session.add(product)
            session.flush()
            run = UpdateRun(
                operator_id=admin.id,
                cutoff_date=date(2026, 7, 17),
                status="processing",
            )
            session.add(run)
            session.flush()
            item = RunItem(run_id=run.id, excel_row=2, original_values={"product_name": "产品A"})
            session.add(item)
            session.commit()
            run_id = run.id
            item_id = item.id
            product_id = product.id
        finally:
            session.close()

        review = client.get(f"/updates/{run_id}/review")
        token = re.search(r'name="token" value="([^"]+)"', review.text).group(1)
        reviewed = client.post(
            f"/updates/{run_id}/items/{item_id}/review",
            data={
                "token": token,
                "product_choice": f"product:{product_id}",
                "weekly": "12.34",
                "review_note": "人工核对",
            },
            follow_redirects=False,
        )

    assert reviewed.status_code == 409
    session = factory()
    try:
        assert session.get(RunItem, item_id).match_source == "none"
    finally:
        session.close()


def test_review_creates_private_product_and_hides_ready_items(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )

    with TestClient(create_app(settings=settings, session_factory=factory)) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        session = factory()
        try:
            admin = session.query(User).filter_by(username="admin").one()
            run = UpdateRun(operator_id=admin.id, cutoff_date=date(2026, 7, 17), status="completed")
            session.add(run)
            session.flush()
            item = RunItem(
                run_id=run.id,
                excel_row=2,
                original_values={"product_name": "测试私募1号"},
                row_status="needs_review",
                metric_values={"weekly": "0.0123"},
                metric_status={"weekly": "extracted", "mtd": "stale"},
            )
            ready_item = RunItem(
                run_id=run.id,
                excel_row=3,
                original_values={"product_name": "完整产品"},
                row_status="ready",
            )
            partial_item = RunItem(
                run_id=run.id,
                excel_row=4,
                original_values={"product_name": "部分产品"},
                row_status="partial",
                metric_values={
                    "weekly": "0.01",
                    "mtd": "0.01",
                    "ytd": "0.01",
                    "annual_2019": "0.01",
                    "annual_2020": "0.01",
                    "annual_2021": "0.01",
                    "annual_2022": "0.01",
                    "annual_2023": "0.01",
                    "annual_2024": "0.01",
                },
                metric_status={"annual_2025": "stale", "sharpe": "stale", "max_drawdown": "stale"},
                error_reason="本次未识别：annual_2025, max_drawdown, sharpe",
            )
            session.add_all([item, ready_item, partial_item])
            session.commit()
            run_id = run.id
            item_id = item.id
        finally:
            session.close()

        preview = client.get(f"/updates/{run_id}/preview")
        assert "识别结果" in preview.text
        assert "待人工审核 1 条" in preview.text
        assert "已确认 1 / 12 项（1 数值＋0 空值）" in preview.text
        assert "已确认 9 / 12 项（9 数值＋0 空值）" in preview.text
        assert "部分识别" in preview.text
        assert "本次未识别" in preview.text
        assert "去审核" in preview.text

        review = client.get(f"/updates/{run_id}/review")
        assert "测试私募1号" in review.text
        assert "完整产品" not in review.text
        assert "部分产品" not in review.text
        assert f'/updates/{run_id}/review?show_all=1' in review.text
        assert 'value="create_private" selected' in review.text
        assert 'class="metric-field missing"' in review.text
        assert "需补录（11 项）" in review.text
        assert "已识别（1 项，可修改）" in review.text
        all_items = client.get(f"/updates/{run_id}/review?show_all=1")
        assert "完整产品" in all_items.text
        assert "部分产品" in all_items.text
        token = re.search(r'name="token" value="([^"]+)"', review.text).group(1)
        saved = client.post(
            f"/updates/{run_id}/items/{item_id}/review",
            data={
                "token": token,
                "product_choice": "create_private",
                "weekly": "1.23",
                "mtd": "2.34",
                "review_note": "核对管理人净值表",
            },
            follow_redirects=False,
        )
        assert saved.status_code == 303

    session = factory()
    try:
        product = session.query(Product).filter_by(product_name="测试私募1号").one()
        assert product.product_type == "private"
        assert session.get(RunItem, item_id).product_id == product.id
        assert session.query(AuditLog).filter_by(action="create_private_product").count() == 1
        assert session.query(AuditLog).filter_by(action="manual_review").count() == 1
    finally:
        session.close()


def test_review_keeps_submitted_values_after_private_code_conflict(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )

    with TestClient(create_app(settings=settings, session_factory=factory)) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        session = factory()
        try:
            admin = session.query(User).filter_by(username="admin").one()
            product_name = "测试私募冲突"
            session.add(
                Product(
                    product_name="其他产品",
                    product_code=private_product_code(product_name),
                    product_type="private",
                )
            )
            run = UpdateRun(operator_id=admin.id, cutoff_date=date(2026, 7, 17), status="completed")
            session.add(run)
            session.flush()
            item = RunItem(
                run_id=run.id,
                excel_row=2,
                original_values={"product_name": product_name},
                row_status="needs_review",
            )
            session.add(item)
            session.commit()
            run_id = run.id
            item_id = item.id
        finally:
            session.close()

        review = client.get(f"/updates/{run_id}/review")
        token = re.search(r'name="token" value="([^"]+)"', review.text).group(1)
        failed = client.post(
            f"/updates/{run_id}/items/{item_id}/review",
            data={
                "token": token,
                "product_choice": "create_private",
                "weekly": "1.23",
                "review_note": "保留这段说明",
            },
        )

    assert failed.status_code == 422
    assert 'value="1.23"' in failed.text
    assert "保留这段说明" in failed.text
    assert 'value="create_private" selected' in failed.text


def test_review_keeps_draft_after_empty_product_choice(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )

    with TestClient(create_app(settings=settings, session_factory=factory)) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        session = factory()
        try:
            admin = session.query(User).filter_by(username="admin").one()
            run = UpdateRun(operator_id=admin.id, cutoff_date=date(2026, 7, 17), status="completed")
            session.add(run)
            session.flush()
            item = RunItem(
                run_id=run.id,
                excel_row=2,
                original_values={"product_name": "测试空选择"},
                row_status="needs_review",
            )
            session.add(item)
            session.commit()
            run_id = run.id
            item_id = item.id
        finally:
            session.close()

        review = client.get(f"/updates/{run_id}/review")
        token = re.search(r'name="token" value="([^"]+)"', review.text).group(1)
        failed = client.post(
            f"/updates/{run_id}/items/{item_id}/review",
            data={
                "token": token,
                "product_choice": "",
                "weekly": "1.23",
                "review_note": "保留这段说明",
            },
        )

    assert failed.status_code == 422
    assert "请选择有效产品" in failed.text
    assert 'value="1.23"' in failed.text
    assert "保留这段说明" in failed.text


def test_review_does_not_create_private_product_for_invalid_metric(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )

    with TestClient(create_app(settings=settings, session_factory=factory)) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        session = factory()
        try:
            admin = session.query(User).filter_by(username="admin").one()
            run = UpdateRun(operator_id=admin.id, cutoff_date=date(2026, 7, 17), status="completed")
            session.add(run)
            session.flush()
            item = RunItem(
                run_id=run.id,
                excel_row=2,
                original_values={"product_name": "测试无效指标"},
                row_status="needs_review",
            )
            session.add(item)
            session.commit()
            run_id = run.id
            item_id = item.id
        finally:
            session.close()

        review = client.get(f"/updates/{run_id}/review")
        token = re.search(r'name="token" value="([^"]+)"', review.text).group(1)
        failed = client.post(
            f"/updates/{run_id}/items/{item_id}/review",
            data={
                "token": token,
                "product_choice": "create_private",
                "weekly": "not-a-number",
                "review_note": "保留这段说明",
            },
        )

    assert failed.status_code == 422
    assert 'value="not-a-number"' in failed.text
    assert 'value="create_private" selected' in failed.text
    session = factory()
    try:
        assert session.query(Product).filter_by(product_name="测试无效指标").count() == 0
    finally:
        session.close()


def test_user_can_manually_review_and_regenerate_a_run(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )
    app = create_app(settings=settings, session_factory=factory)
    with TestClient(app) as client:
        login_page = client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        logged_in = client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )
        assert logged_in.status_code == 303

        catalog_page = client.get("/catalog")
        token = re.search(r'name="token" value="([^"]+)"', catalog_page.text).group(1)
        catalog_response = client.post(
            "/catalog/import",
            data={"token": token},
            files={
                "catalog_file": (
                    "catalog.csv",
                    "product_name,product_code,product_type\n"
                    "仁桥金选泽源5B,P001,private\n"
                    "浑瑾岳桐金选1号B,P002,private\n"
                    "开思金选港股通1号B,P003,private\n"
                    "静瑞金选价值灵动1号B,P004,private\n"
                    "勤辰金选创赢成长1号B,P005,private\n"
                    "易方达环保主题灵活配置混合A,P006,private\n",
                    "text/csv",
                )
            },
            follow_redirects=False,
        )
        assert catalog_response.status_code == 303

        new_page = client.get("/updates/new")
        token = re.search(r'name="token" value="([^"]+)"', new_page.text).group(1)
        created = client.post(
            "/updates/new",
            data={"token": token, "cutoff_date": "2026-07-17"},
            files={
                "workbook": (
                    "template.xlsx",
                    Path("tests/fixtures/net_value_template.xlsx").read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=False,
        )
        assert created.status_code == 303
        run_id = re.search(r"/updates/(\d+)/preview", created.headers["location"]).group(1)

        review = client.get(f"/updates/{run_id}/review")
        assert review.status_code == 200
        assert "人工审核" in review.text
        token = re.search(r'name="token" value="([^"]+)"', review.text).group(1)
        product_id = re.search(r'<option value="product:(\d+)"[^>]*>P001', review.text).group(1)
        item_id = re.search(rf'action="/updates/{run_id}/items/(\d+)/review"', review.text).group(1)

        reviewed = client.post(
            f"/updates/{run_id}/items/{item_id}/review",
            data={
                "token": token,
                "product_choice": f"product:{product_id}",
                "weekly": "12.34",
                "review_note": "人工核对管理人净值表",
            },
            follow_redirects=False,
        )
        assert reviewed.status_code == 303
        preview = client.get(f"/updates/{run_id}/preview")
        assert ">人工审核</td>" in preview.text
        token = re.search(r'name="token" value="([^"]+)"', preview.text).group(1)

        processed = client.post(
            f"/updates/{run_id}/process",
            data={"token": token},
            follow_redirects=False,
        )
        assert processed.status_code == 303
        downloaded = client.get(f"/updates/{run_id}/download")
        assert downloaded.status_code == 404

    session = factory()
    try:
        assert session.query(AuditLog).filter_by(action="manual_review").count() == 1
    finally:
        session.close()


def test_admin_imports_and_user_updates_meeting_record(tmp_path: Path) -> None:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    settings = Settings(
        database_url="sqlite+pysqlite:///:memory:",
        data_dir=tmp_path,
        session_secret="test-secret",
        initial_admin_username="admin",
        initial_admin_password="change-me",
    )
    app = create_app(settings=settings, session_factory=factory)
    with TestClient(app) as admin_client:
        login_page = admin_client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        admin_client.post(
            "/login",
            data={"username": "admin", "password": "change-me", "token": token},
            follow_redirects=False,
        )

        meetings = admin_client.get("/meetings")
        assert meetings.status_code == 200
        token = re.search(r'name="token" value="([^"]+)"', meetings.text).group(1)
        uploaded = admin_client.post(
            "/meetings/import",
            data={"token": token},
            files={
                "workbook": (
                    "meetings.xlsx",
                    meeting_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=False,
        )
        assert uploaded.status_code == 303
        assert "2026陆家嘴论坛" in admin_client.get("/meetings").text

        users = admin_client.get("/admin/users")
        token = re.search(r'name="token" value="([^"]+)"', users.text).group(1)
        created = admin_client.post(
            "/admin/users",
            data={
                "token": token,
                "username": "researcher",
                "password": "user-pass",
                "role": "user",
            },
            follow_redirects=False,
        )
        assert created.status_code == 303

    with TestClient(app) as user_client:
        login_page = user_client.get("/login")
        token = re.search(r'name="token" value="([^"]+)"', login_page.text).group(1)
        user_client.post(
            "/login",
            data={"username": "researcher", "password": "user-pass", "token": token},
            follow_redirects=False,
        )

        detail = user_client.get("/meetings/1")
        assert detail.status_code == 200
        token = re.search(r'name="token" value="([^"]+)"', detail.text).group(1)
        saved = user_client.post(
            "/meetings/1/record",
            data={
                "token": token,
                "company_tags": "券商, 创投",
                "industry_tags": "金融",
                "attendance_status": "attended",
                "minutes": "已参会，关注投融资改革",
                "todo": "跟踪细则",
                "conclusion": "长期利好",
            },
            follow_redirects=False,
        )
        assert saved.status_code == 303
        assert "已参会" in user_client.get("/meetings/1").text
        assert "2026陆家嘴论坛" in user_client.get("/meetings?company=%E5%88%B8%E5%95%86").text

        invalid = user_client.post(
            "/meetings/1/record",
            data={"token": token, "attendance_status": "invalid"},
            follow_redirects=False,
        )
        assert invalid.status_code == 422

        denied = user_client.post("/meetings/import", data={"token": token}, follow_redirects=False)
        assert denied.status_code == 403

    session = factory()
    try:
        assert session.query(Meeting).count() == 1
        imported = session.query(AuditLog).filter_by(
            action="import", object_type="meeting_workbook"
        )
        updated = session.query(AuditLog).filter_by(action="update", object_type="meeting")
        assert imported.count() == 1
        assert updated.count() == 1
    finally:
        session.close()
