from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import Meeting, utcnow

REQUIRED_HEADERS = (
    "会议/事件",
    "日期",
    "性质/层级",
    "核心表述",
    "资本市场影响",
    "投研映射",
    "后续跟踪",
    "来源链接",
    "更新时间",
)
SHEET_NAME = "近期会议更新"


class MeetingImportError(ValueError):
    pass


@dataclass(frozen=True)
class MeetingRow:
    title: str
    date_raw: str
    date_start: date | None
    date_end: date | None
    date_parse_status: str
    level: str
    core_statement: str
    market_impact: str
    research_mapping: str
    follow_up: str
    source_link: str
    source_updated_at: str


@dataclass(frozen=True)
class ImportResult:
    created: int
    updated: int
    skipped: int


def parse_date_range(value: object) -> tuple[date | None, date | None]:
    if isinstance(value, datetime):
        parsed = value.date()
        return parsed, parsed
    if isinstance(value, date):
        return value, value
    if not isinstance(value, str):
        return None, None

    parts = value.strip().split("至")
    if len(parts) not in {1, 2}:
        return None, None
    try:
        start = date.fromisoformat(parts[0].strip())
        end = date.fromisoformat(parts[-1].strip())
    except ValueError:
        return None, None
    if start > end:
        return None, None
    return start, end


def source_key(title: str, date_raw: str) -> str:
    normalized_title = "".join(title.split()).casefold()
    return hashlib.sha256(f"{normalized_title}\n{date_raw.strip()}".encode()).hexdigest()


def read_meeting_rows(path: Path) -> list[MeetingRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise MeetingImportError(f"未找到工作表：{SHEET_NAME}")
        sheet = workbook[SHEET_NAME]
        header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        headers = {_text(value): index for index, value in enumerate(header_row)}
        missing = [header for header in REQUIRED_HEADERS if header not in headers]
        if missing:
            raise MeetingImportError(f"缺少列：{'、'.join(missing)}")

        rows: list[MeetingRow] = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            values = {header: _text(row[index]) for header, index in headers.items()}
            title = values["会议/事件"]
            if not title:
                continue
            start, end = parse_date_range(values["日期"])
            rows.append(
                MeetingRow(
                    title=title,
                    date_raw=values["日期"],
                    date_start=start,
                    date_end=end,
                    date_parse_status="normalized" if start else "unparsed",
                    level=values["性质/层级"],
                    core_statement=values["核心表述"],
                    market_impact=values["资本市场影响"],
                    research_mapping=values["投研映射"],
                    follow_up=values["后续跟踪"],
                    source_link=values["来源链接"],
                    source_updated_at=values["更新时间"],
                )
            )
        return rows
    finally:
        workbook.close()


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def import_meetings(session: Session, path: Path) -> ImportResult:
    rows = read_meeting_rows(path)
    created = 0
    updated = 0
    for row in rows:
        key = source_key(row.title, row.date_raw)
        meeting = session.scalar(select(Meeting).where(Meeting.source_key == key))
        if meeting is None:
            session.add(Meeting(source_key=key, **_source_values(row)))
            created += 1
            continue
        for field, value in _source_values(row).items():
            setattr(meeting, field, value)
        meeting.imported_at = utcnow()
        updated += 1
    return ImportResult(created=created, updated=updated, skipped=0)


def _source_values(row: MeetingRow) -> dict[str, object]:
    return {
        "title": row.title,
        "date_raw": row.date_raw,
        "date_start": row.date_start,
        "date_end": row.date_end,
        "date_parse_status": row.date_parse_status,
        "level": row.level,
        "core_statement": row.core_statement,
        "market_impact": row.market_impact,
        "research_mapping": row.research_mapping,
        "follow_up": row.follow_up,
        "source_link": row.source_link,
        "source_updated_at": row.source_updated_at,
    }
