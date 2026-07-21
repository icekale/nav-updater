from __future__ import annotations

from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from .jobs.review import METRIC_FIELDS
from .models import Product, RunItem, UpdateRun

COMPLETED_RUN_STATUSES = {"completed", "completed_with_warnings"}
MISSING_ITEM_STATUSES = {"needs_review", "stale", "failed"}
MISSING_METRIC_STATUSES = {"stale", "failed", "insufficient_data"}
STATUS_LABELS = {
    "normal": "正常",
    "never_updated": "从未更新",
    "missing_data": "数据缺失",
    "outdated": "更新过期",
}
STATUS_ORDER = {"never_updated": 0, "missing_data": 1, "outdated": 2, "normal": 3}


@dataclass(frozen=True)
class MonitoringRow:
    product_id: int
    product_name: str
    product_code: str
    cutoff_date: date | None
    processed_at: datetime | None
    metric_values: Mapping[str, Decimal | None]
    missing_metrics: tuple[str, ...]
    status: str
    status_label: str
    reason: str
    run_id: int | None
    run_item_id: int | None


@dataclass(frozen=True)
class MonitoringDashboard:
    rows: tuple[MonitoringRow, ...]
    exceptions: tuple[MonitoringRow, ...]
    total_count: int
    normal_count: int
    never_updated_count: int
    missing_data_count: int
    outdated_count: int


def _metric_values(item: RunItem) -> dict[str, Decimal | None]:
    values: dict[str, Decimal | None] = {}
    for field in METRIC_FIELDS:
        raw = item.metric_values.get(field.name)
        if raw is None:
            values[field.name] = None
            continue
        try:
            value = Decimal(str(raw))
        except InvalidOperation:
            values[field.name] = None
            continue
        values[field.name] = value if value.is_finite() else None
    return values


def _missing_metrics(item: RunItem) -> tuple[str, ...]:
    return tuple(
        field.label
        for field in METRIC_FIELDS
        if item.metric_status.get(field.name) in MISSING_METRIC_STATUSES
    )


def _row_for_product(
    product: Product,
    latest: tuple[RunItem, UpdateRun] | None,
    *,
    today: date,
) -> MonitoringRow:
    if latest is None:
        return MonitoringRow(
            product_id=product.id,
            product_name=product.product_name,
            product_code=product.product_code,
            cutoff_date=None,
            processed_at=None,
            metric_values={},
            missing_metrics=(),
            status="never_updated",
            status_label=STATUS_LABELS["never_updated"],
            reason="没有已完成的更新记录",
            run_id=None,
            run_item_id=None,
        )

    item, run = latest
    missing_metrics = _missing_metrics(item)
    metric_values = _metric_values(item)
    if item.row_status in MISSING_ITEM_STATUSES or missing_metrics:
        status = "missing_data"
        reason = (
            f"缺失指标：{'、'.join(missing_metrics)}"
            if missing_metrics
            else "该产品记录仍待人工处理"
        )
    elif run.cutoff_date < today - timedelta(days=10):
        status = "outdated"
        reason = "最近净值更新已超过 10 天"
    else:
        status = "normal"
        reason = "数据完整"
    return MonitoringRow(
        product_id=product.id,
        product_name=product.product_name,
        product_code=product.product_code,
        cutoff_date=run.cutoff_date,
        processed_at=run.finished_at or run.created_at,
        metric_values=metric_values,
        missing_metrics=missing_metrics,
        status=status,
        status_label=STATUS_LABELS[status],
        reason=reason,
        run_id=run.id,
        run_item_id=item.id,
    )


def _sort_rows(rows: list[MonitoringRow]) -> list[MonitoringRow]:
    return sorted(
        rows,
        key=lambda row: (
            STATUS_ORDER[row.status],
            row.cutoff_date or date.min,
            row.product_name,
            row.product_code,
        ),
    )


def build_monitoring_dashboard(
    session: Session,
    *,
    today: date | None = None,
    search: str = "",
    status_filter: str = "all",
) -> MonitoringDashboard:
    today = today or date.today()
    products = session.scalars(
        select(Product)
        .where(Product.is_active.is_(True), Product.product_type == "private")
        .order_by(Product.product_code)
    ).all()
    product_ids = {product.id for product in products}
    latest_by_product: dict[int, tuple[RunItem, UpdateRun]] = {}
    if product_ids:
        completed_items = session.execute(
            select(RunItem, UpdateRun)
            .join(UpdateRun, RunItem.run_id == UpdateRun.id)
            .where(
                RunItem.product_id.in_(product_ids),
                UpdateRun.status.in_(COMPLETED_RUN_STATUSES),
            )
            .order_by(UpdateRun.cutoff_date.desc(), UpdateRun.created_at.desc(), RunItem.id.desc())
        ).all()
        for item, run in completed_items:
            latest_by_product.setdefault(item.product_id, (item, run))

    all_rows = _sort_rows(
        [
            _row_for_product(product, latest_by_product.get(product.id), today=today)
            for product in products
        ]
    )
    counts = {status: sum(row.status == status for row in all_rows) for status in STATUS_LABELS}
    normalized_search = search.strip().lower()
    normalized_filter = status_filter if status_filter in STATUS_LABELS else "all"
    rows = [
        row
        for row in all_rows
        if (
            not normalized_search
            or normalized_search in row.product_name.lower()
            or normalized_search in row.product_code.lower()
        )
        and (normalized_filter == "all" or row.status == normalized_filter)
    ]
    exceptions = tuple(row for row in rows if row.status != "normal")
    return MonitoringDashboard(
        rows=tuple(rows),
        exceptions=exceptions,
        total_count=len(all_rows),
        normal_count=counts["normal"],
        never_updated_count=counts["never_updated"],
        missing_data_count=counts["missing_data"],
        outdated_count=counts["outdated"],
    )


def _export_metric(value: Decimal | None, *, is_percent: bool) -> Decimal | None:
    if value is None:
        return None
    return value * Decimal("100") if is_percent else value


def monitoring_workbook_bytes(rows: Iterable[MonitoringRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "私募产品监控"
    headers = [
        "产品代码",
        "产品名称",
        "数据状态",
        "异常说明",
        "最近净值截止日",
        "最近处理时间",
        *(field.label for field in METRIC_FIELDS),
        "缺失指标",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append(
            [
                row.product_code,
                row.product_name,
                row.status_label,
                row.reason,
                row.cutoff_date.isoformat() if row.cutoff_date else "",
                row.processed_at.strftime("%Y-%m-%d %H:%M") if row.processed_at else "",
                *[
                    _export_metric(
                        row.metric_values.get(field.name),
                        is_percent=field.is_percent,
                    )
                    for field in METRIC_FIELDS
                ],
                "、".join(row.missing_metrics),
            ]
        )
        for column, _ in enumerate(METRIC_FIELDS, start=7):
            sheet.cell(row=sheet.max_row, column=column).number_format = "0.00"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
